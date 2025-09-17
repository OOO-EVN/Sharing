from aiogram import types
from aiogram.dispatcher.filters import BoundFilter
from config import ADMIN_IDS, ALLOWED_CHAT_IDS, SERVICE_ALIASES, YANDEX_SCOOTER_PATTERN, WOOSH_SCOOTER_PATTERN, JET_SCOOTER_PATTERN, BOLT_SCOOTER_PATTERN, BATCH_QUANTITY_PATTERN, TIMEZONE
from database import db_write_batch, db_fetch_all, db_execute
from collections import defaultdict
import datetime

class IsAdminFilter(BoundFilter):
    async def check(self, message: types.Message) -> bool:
        return message.from_user.id in ADMIN_IDS

class IsAllowedChatFilter(BoundFilter):
    async def check(self, message: types.Message) -> bool:
        if message.chat.type == 'private' and message.from_user.id in ADMIN_IDS:
            return True
        if message.chat.type in ['group', 'supergroup'] and message.chat.id in ALLOWED_CHAT_IDS:
            return True
        return False

async def command_start_handler(message: types.Message):
    allowed_chats_info = ', '.join(map(str, ALLOWED_CHAT_IDS)) if ALLOWED_CHAT_IDS else "не указаны"
    response = (
        f"Привет, {message.from_user.full_name}! Я бот для приёма самокатов.\n\n"
        f"Просто отправь мне номер самоката текстом или фотографию с номером в подписи.\n"
        f"Для пакетного приёма используй формат: `сервис количество`.\n\n"
        f"Я работаю в группах с ID: `{allowed_chats_info}` и в личных сообщениях с администраторами.\n"
        f"Твой ID чата: `{message.chat.id}`"
    )
    await message.answer(response, parse_mode="Markdown")

async def process_scooter_text(message: types.Message, text_to_process: str):
    user = message.from_user
    now_localized_str = datetime.datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")

    records_to_insert = []
    accepted_summary = defaultdict(int)

    text_for_numbers = text_to_process

    batch_matches = BATCH_QUANTITY_PATTERN.findall(text_to_process)
    if batch_matches:
        for service_raw, quantity_str in batch_matches:
            service = SERVICE_ALIASES.get(service_raw.lower())
            try:
                quantity = int(quantity_str)
                if service and 0 < quantity <= 200:
                    for i in range(quantity):
                        placeholder_number = f"{service.upper()}_BATCH_{datetime.datetime.now(TIMEZONE).strftime('%H%M%S%f')}_{i+1}"
                        records_to_insert.append((placeholder_number, service, user.id, user.username, user.full_name, now_localized_str, message.chat.id))
                    accepted_summary[service] += quantity
            except (ValueError, TypeError):
                continue
        text_for_numbers = BATCH_QUANTITY_PATTERN.sub('', text_to_process)

    patterns = {
        "Яндекс": YANDEX_SCOOTER_PATTERN,
        "Whoosh": WOOSH_SCOOTER_PATTERN,
        "Jet": JET_SCOOTER_PATTERN,
        "Bolt": BOLT_SCOOTER_PATTERN
    }

    processed_numbers = set()

    for service, pattern in patterns.items():
        numbers = pattern.findall(text_for_numbers)
        for num in numbers:
            raw_num = num.replace('-', '')
            clean_num = raw_num.upper()

            if clean_num in processed_numbers:
                continue

            records_to_insert.append((clean_num, service, user.id, user.username, user.full_name, now_localized_str, message.chat.id))
            accepted_summary[service] += 1
            processed_numbers.add(clean_num)

    if not records_to_insert:
        return False

    await db_write_batch(records_to_insert)

    response_parts = []
    user_mention = f"<a href='tg://user?id={user.id}'>{user.full_name}</a>"
    total_accepted = sum(accepted_summary.values())
    response_parts.append(f"{user_mention}, принято {total_accepted} шт.:")

    for service, count in sorted(accepted_summary.items()):
        if count > 0:
            response_parts.append(f"  - <b>{service}</b>: {count} шт.")

    await message.reply("\n".join(response_parts), parse_mode="HTML")
    return True

async def handle_text_messages(message: types.Message):
    if message.text.startswith('/'):
        return
    await process_scooter_text(message, message.text)

async def handle_photo_messages(message: types.Message):
    if message.caption:
        await process_scooter_text(message, message.caption)

async def handle_unsupported_content(message: types.Message):
    if message.text and message.text.startswith('/'):
        return
    if not (message.photo or (message.text and not message.text.startswith('/'))):
        await message.reply("Извините, я могу обрабатывать только текстовые сообщения и фотографии (с подписями).")

async def today_stats_handler(message: types.Message):
    start_time, end_time, shift_name = get_shift_time_range()

    start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
    end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")

    query = "SELECT service, accepted_by_user_id, accepted_by_username, accepted_by_fullname FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
    records = await db_fetch_all(query, (start_str, end_str))

    if not records:
        await message.answer(f"За {shift_name} пока ничего не принято.")
        return

    user_stats = defaultdict(lambda: defaultdict(int))
    user_info = {}
    service_totals = defaultdict(int)

    for service, user_id, username, fullname in records:
        user_stats[user_id][service] += 1
        service_totals[service] += 1
        if user_id not in user_info:
            user_info[user_id] = f"@{username}" if username else fullname

    response_parts = [f"<b>Статистика за {shift_name} ({start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}):</b>"]
    total_all_users = 0

    for user_id, services in user_stats.items():
        user_total = sum(services.values())
        total_all_users += user_total
        response_parts.append(f"\n<b>{user_info[user_id]}</b> - всего: {user_total} шт.")
        for service, count in sorted(services.items()):
            response_parts.append(f"  - {service}: {count} шт.")

    response_parts.append("\n<b>Итог по сервисам:</b>")
    for service, count in sorted(service_totals.items()):
        response_parts.append(f"<b>{service}</b>: {count} шт.")

    response_parts.append(f"\n<b>Общий итог за {shift_name}: {total_all_users} шт.</b>")

    MESSAGE_LIMIT = 4000
    current_message_buffer = []

    for part in response_parts:
        if len('\n'.join(current_message_buffer)) + len(part) + 1 > MESSAGE_LIMIT:
            if current_message_buffer:
                await message.answer("\n".join(current_message_buffer), parse_mode="HTML")
                current_message_buffer = []
        current_message_buffer.append(part)

    if current_message_buffer:
        await message.answer("\n".join(current_message_buffer), parse_mode="HTML")

def get_shift_time_range():
    now = datetime.datetime.now(TIMEZONE)
    today = now.date()

    morning_shift_start = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(7, 0, 0)))
    morning_shift_end = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
    evening_shift_start = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
    evening_shift_end = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(23, 0, 0)))

    if morning_shift_start <= now < morning_shift_end:
        return morning_shift_start, morning_shift_end, "утреннюю смену"
    elif evening_shift_start <= now < evening_shift_end:
        return evening_shift_start, evening_shift_end, "вечернюю смену"
    else:
        prev_day = today - datetime.timedelta(days=1)
        night_cutoff_current_day = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(4, 0, 0)))
        if TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(0,0,0))) <= now < night_cutoff_current_day:
            prev_evening_shift_start = TIMEZONE.localize(datetime.datetime.combine(prev_day, datetime.time(15, 0, 0)))
            return prev_evening_shift_start, night_cutoff_current_day, "вечернюю смену (с учетом ночных часов)"
        else:
            if now.hour >= 23:
                return evening_shift_start, evening_shift_end, "вечернюю смену"
            return morning_shift_start, morning_shift_end, "утреннюю смену (еще не началась)"

async def export_excel_handler(message: types.Message):
    if not await IsAdminFilter().check(message):
        return

    is_today_shift = message.get_command() == '/export_today_excel'

    await message.answer(f"Формирую отчет...")

    query = "SELECT id, scooter_number, service, accepted_by_user_id, accepted_by_username, accepted_by_fullname, timestamp, chat_id FROM accepted_scooters"

    if is_today_shift:
        start_time, end_time, shift_name = get_shift_time_range()
        start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
        end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        query += " WHERE timestamp BETWEEN ? AND ?"
        records = await db_fetch_all(query, (start_str, end_str))
        date_filter_text = f" за {shift_name}"
    else:
        query += " ORDER BY timestamp DESC"
        records = await db_fetch_all(query)
        date_filter_text = " за все время"

    if not records:
        await message.answer(f"Нет данных для экспорта{date_filter_text}.")
        return

    try:
        excel_file = create_excel_report(records)
        report_type = "shift" if is_today_shift else "full"
        filename = f"report_{report_type}_{datetime.date.today().isoformat()}.xlsx"
        await bot.send_document(message.chat.id, types.InputFile(excel_file, filename=filename), caption=f"Ваш отчет{date_filter_text} готов.")
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel файла: {e}")
        await message.answer("Произошла ошибка при отправке отчета.")

async def service_report_handler(message: types.Message):
    if not await IsAdminFilter().check(message):
        return

    args = message.get_args().split()
    if len(args) != 2:
        await message.reply(
            "Используйте: /service_report <начало> <конец>\nПример: /service_report 2024-07-15 2024-07-25",
            parse_mode=None
        )
        return

    start_date_str, end_date_str = args
    try:
        start_date = TIMEZONE.localize(datetime.datetime.strptime(start_date_str, "%Y-%m-%d"))
        end_date = TIMEZONE.localize(datetime.datetime.strptime(end_date_str, "%Y-%m-%d") + datetime.timedelta(days=1))
    except Exception as e:
        await message.reply(
            "Некорректный формат даты. Дата должна быть в YYYY-MM-DD.",
            parse_mode=None
        )
        return

    report_lines = []
    total_all = 0
    total_service = defaultdict(int)

    current_date = start_date
    while current_date <= end_date:
        morning_start = TIMEZONE.localize(datetime.datetime.combine(current_date.date(), datetime.time(7, 0, 0)))
        morning_end = TIMEZONE.localize(datetime.datetime.combine(current_date.date(), datetime.time(15, 0, 0)))

        morning_query = "SELECT service FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
        morning_records = await db_fetch_all(morning_query, (morning_start.strftime("%Y-%m-%d %H:%M:%S"), morning_end.strftime("%Y-%m-%d %H:%M:%S")))

        morning_services = defaultdict(int)
        morning_total = 0
        for (service,) in morning_records:
            morning_services[service] += 1
            total_service[service] += 1
            morning_total += 1
            total_all += 1

        evening_start = TIMEZONE.localize(datetime.datetime.combine(current_date.date(), datetime.time(15, 0, 0)))
        evening_end = TIMEZONE.localize(datetime.datetime.combine(current_date.date() + datetime.timedelta(days=1), datetime.time(4, 0, 0)))

        evening_query = "SELECT service FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
        evening_records = await db_fetch_all(evening_query, (evening_start.strftime("%Y-%m-%d %H:%M:%S"), evening_end.strftime("%Y-%m-%d %H:%M:%S")))

        evening_services = defaultdict(int)
        evening_total = 0
        for (service,) in evening_records:
            evening_services[service] += 1
            total_service[service] += 1
            evening_total += 1
            total_all += 1

        date_str = current_date.strftime("%d.%m")
        report_lines.append(f"<b>{date_str}</b>")
        report_lines.append("Утренняя смена (7:00-15:00):")
        for service, count in sorted(morning_services.items()):
            report_lines.append(f"{service}: {count} шт.")
        report_lines.append("Вечерняя смена (15:00-4:00):")
        for service, count in sorted(evening_services.items()):
            report_lines.append(f"{service}: {count} шт.")
        day_total = morning_total + evening_total
        report_lines.append(f"<b>Итог за день: {day_total}</b>")
        report_lines.append("")

        current_date += datetime.timedelta(days=1)

    report_lines.append("<b>Итог по сервисам за период:</b>")
    for service, count in sorted(total_service.items()):
        report_lines.append(f"{service}: {count} шт.")
    report_lines.append(f"\n<b>Общий итог: {total_all} шт.</b>")

    report_text = '\n'.join(report_lines)
    MESSAGE_LIMIT = 4000
    buffer = []
    for line in report_lines:
        if len('\n'.join(buffer + [line])) > MESSAGE_LIMIT:
            await message.answer('\n'.join(buffer), parse_mode="HTML")
            buffer = []
        buffer.append(line)

    if buffer:
        await message.answer('\n'.join(buffer), parse_mode="HTML")

def create_excel_report(records: list[tuple]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()
    ws_all_data = wb.active
    ws_all_data.title = "Все данные"

    headers_all_data = ["ID", "Номер Самоката", "Сервис", "ID Пользователя", "Ник", "Полное имя", "Время Принятия", "ID Чата"]
    ws_all_data.append(headers_all_data)
    header_font = Font(bold=True)
    for cell in ws_all_data[1]:
        cell.font = header_font

    for row in records:
        ws_all_data.append(row)

    for col_idx, col in enumerate(ws_all_data.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_all_data.column_dimensions[column_letter].width = adjusted_width

    ws_totals = wb.create_sheet("Итоги")
    totals_headers = ["Пользователь", "Всего Самокатов"]
    ws_totals.append(totals_headers)
    for cell in ws_totals[1]:
        cell.font = header_font

    user_total_counts_summary = defaultdict(int)
    user_info_map_summary = {}

    for record in records:
        user_id = record[3]
        username = record[4]
        fullname = record[5]
        display_name = fullname if fullname else (f"@{username}" if username else f"ID: {user_id}")
        user_total_counts_summary[user_id] += 1
        user_info_map_summary[user_id] = display_name

    sorted_user_ids_summary = sorted(user_total_counts_summary.keys(), key=lambda user_id: user_info_map_summary[user_id].lower())

    for user_id in sorted_user_ids_summary:
        user_display_name = user_info_map_summary[user_id]
        total_count = user_total_counts_summary[user_id]
        ws_totals.append([user_display_name, total_count])
        ws_totals.cell(row=ws_totals.max_row, column=1).font = Font(bold=True)
        ws_totals.cell(row=ws_totals.max_row, column=2).font = Font(bold=True)

    for col_idx, col in enumerate(ws_totals.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_totals.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# Замените все MarkdownV2 на HTML
async def find_scooter_handler(message: types.Message):
    if not await IsAdminFilter().check(message):
        return

    args = message.get_args().strip()
    if not args:
        await message.reply(
            "🔍 Используйте: /find_scooter <номер>\n"
            "Пример:\n"
            "/find_scooter 12345678\n"
            "/find_scooter AB1234"
        )
        return

    scooter_number = args.upper().replace('-', '')

    query = """
        SELECT 
            scooter_number, service, accepted_by_username, accepted_by_fullname, 
            timestamp, chat_id 
        FROM accepted_scooters 
        WHERE scooter_number = ?
        ORDER BY timestamp DESC
    """
    records = await db_fetch_all(query, (scooter_number,))

    if not records:
        await message.reply(f"❌ Нет записей с номером <code>{scooter_number}</code>")
        return

    response_parts = [f"🔍 <b>История самоката <code>{scooter_number}</code>:</b>"]
    for row in records:
        num, service, username, fullname, ts, chat_id = row
        dt = datetime.datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        formatted_time = dt.strftime("%d.%m %H:%M")
        user_display = f"@{username}" if username else fullname
        chat_link = f"<a href='tg://resolve?domain=chat&post={chat_id}'>{chat_id}</a>" if chat_id > 0 else str(chat_id)
        response_parts.append(f"• {service} — {user_display} ({formatted_time}) — чат: {chat_link}")

    MESSAGE_LIMIT = 4000
    current_msg = []
    for line in response_parts:
        if len('\n'.join(current_msg)) + len(line) + 1 > MESSAGE_LIMIT:
            await message.answer('\n'.join(current_msg), parse_mode="HTML")
            current_msg = []
        current_msg.append(line)

    if current_msg:
        await message.answer('\n'.join(current_msg), parse_mode="HTML")


async def delete_scooter_handler(message: types.Message):
    if not await IsAdminFilter().check(message):
        return

    args = message.get_args().split()
    if len(args) < 2:
        await message.reply(
            "🗑️ Используйте: /delete_scooter <номер> <username>\n"
            "Пример:\n"
            "/delete_scooter 12345678 @whoosh_master\n"
            "/delete_scooter AB1234 nobody",
            parse_mode=None  # Убираем разметку для этого сообщения
        )
        return

    scooter_number = args[0].upper().replace('-', '')
    target_username = args[1].lstrip('@')

    if not target_username:
        await message.reply("❌ Укажите корректный username (например, @user или user)", parse_mode=None)
        return

    query = "DELETE FROM accepted_scooters WHERE scooter_number = ? AND accepted_by_username = ?"
    deleted_rows = await db_execute(query, (scooter_number, target_username))

    if deleted_rows > 0:
        await message.reply(
            f"✅ Удалено {deleted_rows} записей:\n"
            f"`{scooter_number}` от пользователя `{target_username}`",
            parse_mode="Markdown"  # Используем Markdown для обратных кавычек
        )
    else:
        await message.reply(
            f"❌ Запись `{scooter_number}` от пользователя `@{target_username}` не найдена.",
            parse_mode="Markdown"  # Используем Markdown для обратных кавычек
        )
