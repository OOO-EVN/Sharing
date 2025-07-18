import asyncio
import re
import os
import sqlite3
import datetime
from io import BytesIO
import pytz
import logging
from concurrent.futures import ThreadPoolExecutor
from collections import defaultdict
from typing import List, Tuple
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# Изменения для aiogram 2.x
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.utils import executor
from aiogram.dispatcher.filters import BoundFilter
from aiogram.utils.exceptions import MessageIsTooLong

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

load_dotenv()

BOT_TOKEN = os.getenv('BOT_TOKEN')
if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN не найден в .env файле. Пожалуйста, добавьте его.")

try:
    ADMIN_IDS = {int(admin_id) for admin_id in os.getenv('ADMIN_IDS', '').split(',') if admin_id.strip()}
    ALLOWED_CHAT_IDS = {int(chat_id) for chat_id in os.getenv('ALLOWED_CHAT_IDS', '').split(',') if chat_id.strip()}
    REPORT_CHAT_IDS = {int(chat_id) for chat_id in os.getenv('REPORT_CHAT_IDS', '').split(',') if chat_id.strip()}
except ValueError:
    logging.error("Не удалось прочитать ADMIN_IDS, ALLOWED_CHAT_IDS или REPORT_CHAT_IDS. Убедитесь, что они являются числами, разделенными запятыми.")
    ADMIN_IDS = set()
    ALLOWED_CHAT_IDS = set()
    REPORT_CHAT_IDS = set()

DB_NAME = 'scooters.db'
TIMEZONE = pytz.timezone('Asia/Almaty') # Ваша таймзона UTC+5

YANDEX_SCOOTER_PATTERN = re.compile(r'\b(\d{8})\b')
WOOSH_SCOOTER_PATTERN = re.compile(r'\b([A-ZА-Я]{2}\d{4})\b', re.IGNORECASE)
JET_SCOOTER_PATTERN = re.compile(r'\b(\d{3}-?\d{3})\b')

BATCH_QUANTITY_PATTERN = re.compile(r'\b(whoosh|jet|yandex|вуш|джет|яндекс|w|j|y)\s+(\d+)\b', re.IGNORECASE)
SERVICE_ALIASES = {
    "yandex": "Яндекс", "яндекс": "Яндекс", "y": "Яндекс",
    "whoosh": "Whoosh", "вуш": "Whoosh", "w": "Whoosh",
    "jet": "Jet", "джет": "Jet", "j": "Jet"
}
SERVICE_MAP = {"yandex": "Яндекс", "whoosh": "Whoosh", "jet": "Jet"}

bot = Bot(token=BOT_TOKEN, parse_mode="HTML")
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

db_executor = None
scheduler = None

class IsAdminFilter(BoundFilter):
    async def check(self, message: types.Message) -> bool:
        return message.from_user.id in ADMIN_IDS

class IsAllowedChatFilter(BoundFilter):
    async def check(self, message: types.Message) -> bool:
        if message.chat.type == 'private' and message.from_user.id in ADMIN_IDS:
            return True
        if message.chat.type in ['group', 'supergroup'] and message.chat.id in ALLOWED_CHAT_IDS:
            return True
        logging.warning(f"Сообщение от {message.from_user.id} в чате {message.chat.id} было заблокировано фильтром.")
        return False

def run_db_query(query: str, params: tuple = (), fetch: str = None):
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        conn.execute("PRAGMA journal_mode=WAL;")
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()
        if fetch == 'one':
            return cursor.fetchone()
        if fetch == 'all':
            return cursor.fetchall()
        return cursor.lastrowid
    except sqlite3.Error as e:
        logging.error(f"Ошибка базы данных: {e}\nЗапрос: {query}")
        return None
    finally:
        if conn:
            conn.close()

def init_db():
    run_db_query('''
        CREATE TABLE IF NOT EXISTS accepted_scooters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scooter_number TEXT NOT NULL,
            service TEXT NOT NULL,
            accepted_by_user_id INTEGER NOT NULL,
            accepted_by_username TEXT,
            accepted_by_fullname TEXT NOT NULL,
            timestamp DATETIME NOT NULL,
            chat_id INTEGER NOT NULL
        )
    ''')
    run_db_query("CREATE INDEX IF NOT EXISTS idx_timestamp ON accepted_scooters (timestamp);")
    run_db_query("CREATE INDEX IF NOT EXISTS idx_user_service ON accepted_scooters (accepted_by_user_id, service);")
    logging.info("База данных успешно инициализирована.")

def insert_batch_records(records_data: List[Tuple]):
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        conn.execute("PRAGMA journal_mode=WAL;")
        cursor = conn.cursor()
        cursor.executemany('''
            INSERT INTO accepted_scooters (scooter_number, service, accepted_by_user_id, accepted_by_username, accepted_by_fullname, timestamp, chat_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', records_data)
        conn.commit()
    except sqlite3.Error as e:
        logging.error(f"Ошибка пакетной вставки в БД: {e}")
    finally:
        if conn:
            conn.close()

async def db_write_batch(records_data: List[Tuple]):
    loop = asyncio.get_running_loop()
    global db_executor
    await loop.run_in_executor(db_executor, insert_batch_records, records_data)

async def db_fetch_all(query: str, params: tuple = ()):
    loop = asyncio.get_running_loop()
    global db_executor
    return await loop.run_in_executor(db_executor, run_db_query, query, params, 'all')

@dp.message_handler(IsAllowedChatFilter(), commands="start")
async def command_start_handler(message: types.Message):
    allowed_chats_info = ', '.join(map(str, ALLOWED_CHAT_IDS)) if ALLOWED_CHAT_IDS else "не указаны"
    response = (
        f"Привет, {message.from_user.full_name}! Я бот для приёма самокатов.\n\n"
        f"Просто отправь мне **номер самоката текстом** или **фотографию с номером в подписи**.\n"
        f"Для пакетного приёма используй формат: `сервис количество` (например, `Яндекс 10`, `y 5`, `Whoosh 15`, `w 20`, `Jet 8`, `j 3`).\n\n"
        f"Я работаю в группах с ID: `{allowed_chats_info}` и в личных сообщениях с администраторами.\n"
        f"Твой ID чата: `{message.chat.id}`"
    )
    await message.answer(response, parse_mode="Markdown")

@dp.message_handler(IsAllowedChatFilter(), commands="batch_accept")
async def batch_accept_handler(message: types.Message):
    args = message.get_args().split()
    if len(args) != 2:
        await message.reply("Используйте: `/batch_accept <сервис> <количество>`\nПример: `/batch_accept Yandex 20` или `/batch_accept y 20`", parse_mode="Markdown")
        return

    service_raw, quantity_str = args
    service = SERVICE_ALIASES.get(service_raw.lower())

    if not service:
        await message.reply("Неизвестный сервис. Доступны: `Yandex` (`y`), `Whoosh` (`w`), `Jet` (`j`).", parse_mode="Markdown")
        return
    try:
        quantity = int(quantity_str)
        if not (0 < quantity <= 200):
            raise ValueError
    except ValueError:
        await message.reply("Количество должно быть числом от 1 до 200.")
        return

    user = message.from_user
    now_localized_str = datetime.datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")

    records_to_insert = [
        (
            f"{service.upper()}_BATCH_{i+1}",
            service, user.id, user.username, user.full_name, now_localized_str, message.chat.id
        ) for i in range(quantity)
    ]

    await db_write_batch(records_to_insert)

    user_mention = types.User.get_mention(user)
    await message.reply(f"{user_mention}, принято {quantity} самокатов сервиса <b>{service}</b>.")

def get_shift_time_range_for_report(shift_type: str):
    now = datetime.datetime.now(TIMEZONE)
    today = now.date()
    
    if shift_type == 'morning':
        start_time = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(7, 0, 0)))
        end_time = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
        shift_name = "утреннюю смену"
    elif shift_type == 'evening':
        evening_start_actual = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
        evening_end_extended = TIMEZONE.localize(datetime.datetime.combine(today + datetime.timedelta(days=1), datetime.time(4, 0, 0)))
        
        start_time = evening_start_actual
        end_time = evening_end_extended
        shift_name = "вечернюю смену (с учетом ночных часов)"
    else:
        return None, None, None
        
    return start_time, end_time, shift_name


def get_shift_time_range():
    now = datetime.datetime.now(TIMEZONE)
    today = now.date()

    morning_shift_start = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(7, 0, 0)))
    morning_shift_end = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
    evening_shift_start = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
    evening_shift_end = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(23, 0, 0)))

    if morning_shift_start <= now < morning_shift_end:
        return morning_shift_start, morning_shift_end, "утреннюю смену"
    elif evening_shift_start <= now < evening_shift_end:
        return evening_shift_start, evening_shift_end, "вечернюю смену"
    else:
        prev_day = today - datetime.timedelta(days=1)
        night_cutoff_current_day = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(4, 0, 0)))

        if TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(0,0,0))) <= now < night_cutoff_current_day:
            prev_evening_shift_start = TIMEZONE.localize(datetime.datetime.combine(prev_day, datetime.time(15, 0, 0)))
            return prev_evening_shift_start, night_cutoff_current_day, "вечернюю смену (с учетом ночных часов)"
        else:
            if now.hour >= 23:
                return evening_shift_start, evening_shift_end, "вечернюю смену"
            return morning_shift_start, morning_shift_end, "утреннюю смену (еще не началась)"


@dp.message_handler(IsAdminFilter(), commands="today_stats")
async def today_stats_handler(message: types.Message):
    start_time, end_time, shift_name = get_shift_time_range()
    
    start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
    end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")

    query = "SELECT service, accepted_by_user_id, accepted_by_username, accepted_by_fullname FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
    records = await db_fetch_all(query, (start_str, end_str))

    if not records:
        await message.answer(f"За {shift_name} пока ничего не принято.")
        return

    user_stats = defaultdict(lambda: defaultdict(int))
    user_info = {}
    service_totals = defaultdict(int)

    for service, user_id, username, fullname in records:
        user_stats[user_id][service] += 1
        service_totals[service] += 1
        if user_id not in user_info:
            user_info[user_id] = f"@{username}" if username else fullname

    response_parts = [f"<b>Статистика за {shift_name} ({start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}):</b>"]
    total_all_users = 0

    for user_id, services in user_stats.items():
        user_total = sum(services.values())
        total_all_users += user_total
        response_parts.append(f"\n<b>{user_info[user_id]}</b> - всего: {user_total} шт.")
        for service, count in sorted(services.items()):
            response_parts.append(f"  - {service}: {count} шт.")

    response_parts.append("\n<b>Итог по сервисам:</b>")
    for service, count in sorted(service_totals.items()):
        response_parts.append(f"<b>{service}</b>: {count} шт.")

    response_parts.append(f"\n<b>Общий итог за {shift_name}: {total_all_users} шт.</b>")
    
    MESSAGE_LIMIT = 4000
    current_message_buffer = []
    
    for part in response_parts:
        if len('\n'.join(current_message_buffer)) + len(part) + (1 if current_message_buffer else 0) > MESSAGE_LIMIT:
            if current_message_buffer:
                await message.answer("\n".join(current_message_buffer))
                current_message_buffer = []
        current_message_buffer.append(part)
    
    if current_message_buffer:
        await message.answer("\n".join(current_message_buffer))


@dp.message_handler(IsAdminFilter(), commands=["export_today_excel", "export_all_excel"])
async def export_excel_handler(message: types.Message):
    is_today_shift = message.get_command() == '/export_today_excel'
    
    await message.answer(f"Формирую отчет...")

    query = "SELECT id, scooter_number, service, accepted_by_user_id, accepted_by_username, accepted_by_fullname, timestamp, chat_id FROM accepted_scooters"
    
    if is_today_shift:
        start_time, end_time, shift_name = get_shift_time_range()
        start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
        end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        query += " WHERE timestamp BETWEEN ? AND ?"
        records = await db_fetch_all(query, (start_str, end_str))
        date_filter_text = f" за {shift_name}"
    else:
        query += " ORDER BY timestamp DESC"
        records = await db_fetch_all(query)
        date_filter_text = " за все время"

    if not records:
        await message.answer(f"Нет данных для экспорта{date_filter_text}.")
        logging.info(f"Нет данных для экспорта отчета{date_filter_text}.")
        return

    try:
        excel_file = create_excel_report(records)
        report_type = "shift" if is_today_shift else "full"
        filename = f"report_{report_type}_{datetime.date.today().isoformat()}.xlsx"
        
        logging.info(f"Попытка отправить Excel файл: {filename}, размер: {excel_file.getbuffer().nbytes} байт.")
        await bot.send_document(message.chat.id, types.InputFile(excel_file, filename=filename), caption=f"Ваш отчет{date_filter_text} готов.")
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel файла: {e}", exc_info=True)
        await message.answer("Произошла ошибка при отправке отчета. Пожалуйста, свяжитесь с администратором.")

def create_excel_report(records: List[Tuple]) -> BytesIO:
    wb = Workbook()
    
    # --- ЛИСТ "ВСЕ ДАННЫЕ" ---
    # Этот лист остается для просмотра всех записей подряд
    ws_all_data = wb.active
    ws_all_data.title = "Все данные"

    headers_all_data = ["ID", "Номер Самоката", "Сервис", "ID Пользователя", "Ник", "Полное имя", "Время Принятия", "ID Чата"]
    ws_all_data.append(headers_all_data)
    header_font = Font(bold=True)
    for cell in ws_all_data[1]:
        cell.font = header_font

    for row in records:
        ws_all_data.append(row)

    # Автонастройка ширины столбцов на листе "Все данные"
    for col_idx, col in enumerate(ws_all_data.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_all_data.column_dimensions[column_letter].width = adjusted_width

    # --- ЛИСТ "ИТОГИ" (сводка по всем пользователям) ---
    # Этот лист будет содержать общую сводку по всем пользователям
    ws_totals = wb.create_sheet("Итоги")
    totals_headers = ["Пользователь", "Всего Самокатов"]
    ws_totals.append(totals_headers)
    for cell in ws_totals[1]:
        cell.font = header_font

    user_total_counts_summary = defaultdict(int)
    user_info_map_summary = {} # Для отображаемых имен пользователей

    for record in records:
        user_id = record[3]
        username = record[4]
        fullname = record[5]
        display_name = fullname if fullname else (f"@{username}" if username else f"ID: {user_id}")
        user_total_counts_summary[user_id] += 1
        user_info_map_summary[user_id] = display_name

    sorted_user_ids_summary = sorted(user_total_counts_summary.keys(), key=lambda user_id: user_info_map_summary[user_id].lower())

    for user_id in sorted_user_ids_summary:
        user_display_name = user_info_map_summary[user_id]
        total_count = user_total_counts_summary[user_id]
        ws_totals.append([user_display_name, total_count])
        # Выделяем жирным
        ws_totals.cell(row=ws_totals.max_row, column=1).font = Font(bold=True)
        ws_totals.cell(row=ws_totals.max_row, column=2).font = Font(bold=True)

    # Автонастройка ширины столбцов на листе "Итоги"
    for col_idx, col in enumerate(ws_totals.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_totals.column_dimensions[column_letter].width = adjusted_width

    # --- ОТДЕЛЬНЫЕ ЛИСТЫ ДЛЯ КАЖДОГО ПОЛЬЗОВАТЕЛЯ ---
    user_records = defaultdict(list)
    user_info_for_sheets = {}

    for record in records:
        user_id = record[3]
        username = record[4]
        fullname = record[5]
        display_name = fullname if fullname else (f"@{username}" if username else f"ID: {user_id}")
        
        user_records[user_id].append(record)
        if user_id not in user_info_for_sheets:
            user_info_for_sheets[user_id] = display_name

    sorted_user_ids = sorted(user_records.keys(), key=lambda user_id: user_info_for_sheets[user_id].lower())

    user_sheet_headers = ["ID", "Номер Самоката", "Сервис", "Время Принятия", "ID Чата"] # Заголовки для листов пользователей

    for user_id in sorted_user_ids:
        user_display_name = user_info_for_sheets[user_id]
        # Telegram username может быть длиной до 32 символов.
        # Имя листа в Excel не может превышать 31 символ и не может содержать некоторые спец. символы.
        # Очищаем имя для листа
        sheet_name_raw = f"{user_display_name[:25].replace('@', '')}" # Обрезаем и убираем @
        # Удаляем недопустимые символы для имени листа
        invalid_chars = re.compile(r'[\\/:*?"<>|]')
        sheet_name = invalid_chars.sub('', sheet_name_raw)
        
        # Добавляем префикс ID, если после очистки имя пустое или слишком короткое,
        # или если оно совпадает с уже существующим листом (крайне маловероятно после обработки)
        if not sheet_name or len(sheet_name) < 3: # Если имя слишком короткое после очистки
             sheet_name = f"ID{user_id}"
        
        # Убедимся, что имя листа уникально (если вдруг совпадения по коротким именам)
        original_sheet_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_sheet_name[:28]}{counter}" # Добавляем счетчик, чтобы не превысить 31 символ
            counter += 1

        ws_user = wb.create_sheet(title=sheet_name)
        
        ws_user.append(user_sheet_headers)
        for cell in ws_user[1]:
            cell.font = header_font

        current_user_total = 0
        user_service_breakdown = defaultdict(int) # Для статистики по сервисам на листе пользователя

        # Записи для текущего пользователя
        for record in user_records[user_id]:
            # Отфильтровываем user_id, username, fullname, chat_id (дублируются в названии листа)
            # Структура record: ID, Номер Самоката, Сервис, ID Пользователя, Ник, Полное имя, Время Принятия, ID Чата
            # Новая строка: ID, Номер Самоката, Сервис, Время Принятия, ID Чата
            row_to_add = [record[0], record[1], record[2], record[6], record[7]]
            ws_user.append(row_to_add)
            current_user_total += 1
            user_service_breakdown[record[2]] += 1 # Считаем по сервисам для этого пользователя

        # Добавляем итоговую информацию в конце листа пользователя
        ws_user.append([]) # Пустая строка для разделения
        
        ws_user.append(["Статистика по сервисам:"])
        ws_user.cell(row=ws_user.max_row, column=1).font = Font(bold=True)
        ws_user.merge_cells(start_row=ws_user.max_row, start_column=1, end_row=ws_user.max_row, end_column=2)

        for service, count in sorted(user_service_breakdown.items()):
            ws_user.append([service, count])

        ws_user.append([]) # Еще одна пустая строка
        
        ws_user.append(["Всего принято:", current_user_total])
        ws_user.cell(row=ws_user.max_row, column=1).font = Font(bold=True)
        ws_user.cell(row=ws_user.max_row, column=2).font = Font(bold=True)
        ws_user.cell(row=ws_user.max_row, column=1).alignment = Alignment(horizontal='right')
        ws_user.cell(row=ws_user.max_row, column=2).alignment = Alignment(horizontal='center')

        # Автонастройка ширины столбцов на листе пользователя
        for col_idx, col in enumerate(ws_user.columns):
            max_length = 0
            column_letter = get_column_letter(col_idx + 1)
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_user.column_dimensions[column_letter].width = adjusted_width
            
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

async def process_scooter_text(message: types.Message, text_to_process: str):
    user = message.from_user
    now_localized_str = datetime.datetime.now(TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")

    records_to_insert = []
    accepted_summary = defaultdict(int)
    
    text_for_numbers = text_to_process

    batch_matches = BATCH_QUANTITY_PATTERN.findall(text_to_process)
    if batch_matches:
        for service_raw, quantity_str in batch_matches:
            service = SERVICE_ALIASES.get(service_raw.lower())
            try:
                quantity = int(quantity_str)
                if service and 0 < quantity <= 200:
                    for i in range(quantity):
                        placeholder_number = f"{service.upper()}_BATCH_{datetime.datetime.now().strftime('%H%M%S%f')}_{i+1}" 
                        records_to_insert.append((placeholder_number, service, user.id, user.username, user.full_name, now_localized_str, message.chat.id))
                    accepted_summary[service] += quantity
            except (ValueError, TypeError):
                continue
        text_for_numbers = BATCH_QUANTITY_PATTERN.sub('', text_to_process)

    patterns = {
        "Яндекс": YANDEX_SCOOTER_PATTERN,
        "Whoosh": WOOSH_SCOOTER_PATTERN,
        "Jet": JET_SCOOTER_PATTERN
    }
    
    processed_numbers = set()

    for service, pattern in patterns.items():
        numbers = pattern.findall(text_for_numbers)
        for num in numbers:
            clean_num = num.replace('-', '') if service == "Jet" else num.upper()
            
            if clean_num in processed_numbers:
                continue
            
            records_to_insert.append((clean_num, service, user.id, user.username, user.full_name, now_localized_str, message.chat.id))
            accepted_summary[service] += 1
            processed_numbers.add(clean_num)

    if not records_to_insert:
        return False

    await db_write_batch(records_to_insert)

    response_parts = []
    user_mention = types.User.get_mention(user)
    total_accepted = sum(accepted_summary.values())
    response_parts.append(f"{user_mention}, принято {total_accepted} шт.:")

    for service, count in sorted(accepted_summary.items()):
        if count > 0:
            response_parts.append(f"  - <b>{service}</b>: {count} шт.")

    await message.reply("\n".join(response_parts))
    return True

@dp.message_handler(IsAllowedChatFilter(), content_types=types.ContentTypes.TEXT)
async def handle_text_messages(message: types.Message):
    if message.text.startswith('/'):
        return
    await process_scooter_text(message, message.text)
@dp.message_handler(IsAdminFilter(), commands=["service_report"])
async def service_report_handler(message: types.Message):
    args = message.get_args().split()
    if len(args) != 2:
        await message.reply(
            "Используйте: /service_report <начало> <конец>\n"
            "Пример: /service_report 2024-07-15 2024-07-25"
        )
        return

    start_date_str, end_date_str = args
    try:
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except Exception:
        await message.reply("Некорректный формат даты. Дата должна быть в YYYY-MM-DD.")
        return

    report_lines = []
    total_all = 0
    total_service = defaultdict(int)

    for cur_date in (start_date + datetime.timedelta(n) for n in range((end_date - start_date).days + 1)):
        # Утренняя смена
        morning_start = TIMEZONE.localize(datetime.datetime.combine(cur_date, datetime.time(7, 0, 0)))
        morning_end = TIMEZONE.localize(datetime.datetime.combine(cur_date, datetime.time(15, 0, 0)))
        morning_query = "SELECT service FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
        morning_records = await db_fetch_all(morning_query, (morning_start.strftime("%Y-%m-%d %H:%M:%S"), morning_end.strftime("%Y-%m-%d %H:%M:%S")))
        morning_services = defaultdict(int)
        for (service,) in morning_records:
            morning_services[service] += 1
            total_service[service] += 1
            total_all += 1

        # Вечерняя смена (с учетом ночи)
        even_start = TIMEZONE.localize(datetime.datetime.combine(cur_date, datetime.time(15, 0, 0)))
        even_end = TIMEZONE.localize(datetime.datetime.combine(cur_date + datetime.timedelta(days=1), datetime.time(4, 0, 0)))
        even_query = "SELECT service FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
        even_records = await db_fetch_all(even_query, (even_start.strftime("%Y-%m-%d %H:%M:%S"), even_end.strftime("%Y-%m-%d %H:%M:%S")))
        even_services = defaultdict(int)
        for (service,) in even_records:
            even_services[service] += 1
            total_service[service] += 1
            total_all += 1

        # Формируем строки по дням
        date_str = cur_date.strftime("%d.%m")
        report_lines.append(f"<b>{date_str}</b>")
        report_lines.append("Утренняя смена:")
        for service, count in sorted(morning_services.items()):
            report_lines.append(f"{service}: {count} шт.")
        report_lines.append("Вечерняя смена:")
        for service, count in sorted(even_services.items()):
            report_lines.append(f"{service}: {count} шт.")
        report_lines.append("")  # пустая строка-разделитель

    # Итог по всему периоду
    report_lines.append("<b>Итог по сервисам за период:</b>")
    for service, count in sorted(total_service.items()):
        report_lines.append(f"{service}: {count} шт.")
    report_lines.append(f"\n<b>Общий итог: {total_all} шт.</b>")

    # Отправляем по частям, если превышает лимит
    MESSAGE_LIMIT = 4000
    buffer = []
    for line in report_lines:
        if len('\n'.join(buffer)) + len(line) + 1 > MESSAGE_LIMIT:
            await message.answer('\n'.join(buffer))
            buffer = []
        buffer.append(line)
    if buffer:
        await message.answer('\n'.join(buffer))
@dp.message_handler(IsAllowedChatFilter(), content_types=types.ContentTypes.PHOTO)
async def handle_photo_messages(message: types.Message):
    if message.caption:
        await process_scooter_text(message, message.caption)

@dp.message_handler(IsAllowedChatFilter(), content_types=types.ContentTypes.ANY)
async def handle_unsupported_content(message: types.Message):
    if message.text and message.text.startswith('/'):
        return
    if not (message.photo or (message.text and not message.text.startswith('/'))):
        await message.reply("Извините, я могу обрабатывать только текстовые сообщения и фотографии (с подписями). "
                            "Видео, документы и другие файлы я не поддерживаю.")

async def send_scheduled_report(shift_type: str):
    logging.info(f"Запуск отправки автоматического отчета для {shift_type} смены.")
    
    start_time, end_time, shift_name = get_shift_time_range_for_report(shift_type)
    
    if not start_time or not end_time:
        logging.error(f"Не удалось определить временной диапазон для отчета '{shift_type}' смены.")
        return

    start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
    end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")

    query = "SELECT id, scooter_number, service, accepted_by_user_id, accepted_by_username, accepted_by_fullname, timestamp, chat_id FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
    records = await db_fetch_all(query, (start_str, end_str))

    if not records:
        message_text = f"Отчет за {shift_name} ({start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}): За смену ничего не принято."
        for chat_id in REPORT_CHAT_IDS:
            try:
                await bot.send_message(chat_id, message_text)
                logging.info(f"Отправлено уведомление об отсутствии данных за {shift_name} в чат {chat_id}.")
            except Exception as e:
                logging.error(f"Ошибка отправки уведомления в чат {chat_id}: {e}")
        return

    try:
        excel_file = create_excel_report(records)
        report_type_filename = "morning_shift" if shift_type == 'morning' else "evening_shift"
        filename = f"report_{report_type_filename}_{start_time.strftime('%Y%m%d')}.xlsx"
        caption = f"Ежедневный отчет за {shift_name} ({start_time.strftime('%d.%m %H:%M')} - {end_time.strftime('%d.%m %H:%M')})"
        
        for chat_id in REPORT_CHAT_IDS:
            try:
                excel_file.seek(0) 
                await bot.send_document(chat_id, types.InputFile(excel_file, filename=filename), caption=caption)
                logging.info(f"Отправлен Excel отчет за {shift_name} в чат {chat_id}.")
            except Exception as e:
                logging.error(f"Ошибка отправки Excel файла в чат {chat_id}: {e}", exc_info=True)

    except Exception as e:
        logging.error(f"Произошла общая ошибка при формировании или отправке Excel отчета: {e}", exc_info=True)
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, f"Ошибка при формировании/отправке отчета за {shift_name}: {e}")
            except Exception as err:
                logging.error(f"Не удалось отправить уведомление об ошибке администратору {admin_id}: {err}")


async def on_startup(dispatcher: Dispatcher):
    global db_executor
    db_executor = ThreadPoolExecutor(max_workers=5)
    
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(db_executor, init_db)
    
    global scheduler
    scheduler = AsyncIOScheduler(timezone=str(TIMEZONE))
    
    scheduler.add_job(send_scheduled_report, 'cron', hour=15, minute=0, timezone=str(TIMEZONE), args=['morning'])
    logging.info("Задача для отправки утреннего отчета (в 15:00) запланирована.")
    
    scheduler.add_job(send_scheduled_report, 'cron', hour=23, minute=0, timezone=str(TIMEZONE), args=['evening'])
    logging.info("Задача для отправки вечернего отчета (в 23:00) запланирована.")
    
    scheduler.start()
    logging.info("APScheduler запущен.")
    
    admin_commands = [
        types.BotCommand(command="start", description="Начало работы"),
        types.BotCommand(command="today_stats", description="Статистика за текущую смену"),
        types.BotCommand(command="export_today_excel", description="Экспорт Excel за текущую смену"),
        types.BotCommand(command="export_all_excel", description="Экспорт Excel за все время"),
        types.BotCommand(command="service_report", description="Отчет по сервисам за период"),
        types.BotCommand(command="batch_accept", description="Пакетный прием (сервис кол-во)"),
    ]
    await dispatcher.bot.set_my_commands(admin_commands)
    logging.info("Бот запущен и команды установлены.")


async def on_shutdown(dispatcher: Dispatcher):
    global db_executor
    if db_executor:
        db_executor.shutdown(wait=True)
    
    global scheduler
    if scheduler:
        scheduler.shutdown()
        logging.info("APScheduler остановлен.")
        
    logging.info("Пул потоков БД остановлен.")
    logging.info("Бот остановлен.")

if __name__ == "__main__":
    executor.start_polling(dp, on_startup=on_startup, on_shutdown=on_shutdown, skip_updates=True)
