# reports.py
from aiogram import types
from config import TIMEZONE, REPORT_CHAT_IDS
from database import db_fetch_all
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import logging
from collections import defaultdict
from apscheduler.schedulers.asyncio import AsyncIOScheduler

scheduler = AsyncIOScheduler(timezone=str(TIMEZONE))

def create_excel_report(records: list[tuple]) -> BytesIO:
    wb = Workbook()
    ws_all_data = wb.active
    ws_all_data.title = "Все данные"

    headers_all_data = ["ID", "Номер Самоката", "Сервис", "ID Пользователя", "Ник", "Полное имя", "Время Принятия", "ID Чата"]
    ws_all_data.append(headers_all_data)
    header_font = Font(bold=True)
    for cell in ws_all_data[1]:
        cell.font = header_font

    for row in records:
        ws_all_data.append(row)

    for col_idx, col in enumerate(ws_all_data.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_all_data.column_dimensions[column_letter].width = adjusted_width

    ws_totals = wb.create_sheet("Итоги")
    totals_headers = ["Пользователь", "Всего Самокатов"]
    ws_totals.append(totals_headers)
    for cell in ws_totals[1]:
        cell.font = header_font

    user_total_counts_summary = defaultdict(int)
    user_info_map_summary = {}

    for record in records:
        user_id = record[3]
        username = record[4]
        fullname = record[5]
        display_name = fullname if fullname else (f"@{username}" if username else f"ID: {user_id}")
        user_total_counts_summary[user_id] += 1
        user_info_map_summary[user_id] = display_name

    sorted_user_ids_summary = sorted(user_total_counts_summary.keys(), key=lambda user_id: user_info_map_summary[user_id].lower())

    for user_id in sorted_user_ids_summary:
        user_display_name = user_info_map_summary[user_id]
        total_count = user_total_counts_summary[user_id]
        ws_totals.append([user_display_name, total_count])
        ws_totals.cell(row=ws_totals.max_row, column=1).font = Font(bold=True)
        ws_totals.cell(row=ws_totals.max_row, column=2).font = Font(bold=True)

    for col_idx, col in enumerate(ws_totals.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_totals.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def get_shift_time_range_for_report(shift_type: str):
    now = datetime.datetime.now(TIMEZONE)
    today = now.date()
    
    if shift_type == 'morning':
        start_time = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(7, 0, 0)))
        end_time = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
        shift_name = "утреннюю смену"
    elif shift_type == 'evening':
        evening_start_actual = TIMEZONE.localize(datetime.datetime.combine(today, datetime.time(15, 0, 0)))
        evening_end_extended = TIMEZONE.localize(datetime.datetime.combine(today + datetime.timedelta(days=1), datetime.time(4, 0, 0)))
        start_time = evening_start_actual
        end_time = evening_end_extended
        shift_name = "вечернюю смену (с учетом ночных часов)"
    else:
        return None, None, None
        
    return start_time, end_time, shift_name

# 🔥 ИСПРАВЛЕНА: добавлено логирование и проверка bot_instance
async def send_scheduled_report(shift_type: str, bot_instance):
    if bot_instance is None:
        logging.error("❌ bot_instance is None in send_scheduled_report!")
        return

    start_time, end_time, shift_name = get_shift_time_range_for_report(shift_type)
    if not start_time or not end_time:
        logging.warning(f"Не удалось определить время смены для {shift_type}")
        return

    start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
    end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")

    query = "SELECT id, scooter_number, service, accepted_by_user_id, accepted_by_username, accepted_by_fullname, timestamp, chat_id FROM accepted_scooters WHERE timestamp BETWEEN ? AND ?"
    records = await db_fetch_all(query, (start_str, end_str))

    if not records:
        message_text = f"Отчет за {shift_name} ({start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}): За смену ничего не принято."
        for chat_id in REPORT_CHAT_IDS:
            try:
                await bot_instance.send_message(chat_id, message_text)
                logging.info(f"✅ Текстовый отчёт отправлен в {chat_id}")
            except Exception as e:
                logging.error(f"❌ Не удалось отправить текст в {chat_id}: {e}")
        return

    try:
        excel_file = create_excel_report(records)
        report_type_filename = "morning_shift" if shift_type == 'morning' else "evening_shift"
        filename = f"report_{report_type_filename}_{start_time.strftime('%Y%m%d')}.xlsx"
        caption = f"Ежедневный отчет за {shift_name} ({start_time.strftime('%d.%m %H:%M')} - {end_time.strftime('%d.%m %H:%M')})"
        
        for chat_id in REPORT_CHAT_IDS:
            try:
                excel_file.seek(0)
                await bot_instance.send_document(
                    chat_id,
                    types.InputFile(excel_file, filename=filename),
                    caption=caption
                )
                logging.info(f"✅ Excel-отчёт отправлен в {chat_id}")
            except Exception as e:
                logging.error(f"❌ Ошибка отправки Excel в {chat_id}: {e}")
    except Exception as e:
        logging.error(f"❌ Ошибка генерации Excel-отчёта: {e}")
